"""Excel parser: decrypt and extract standardized data from all detail sheets."""

import io
import re
from dataclasses import dataclass
from pathlib import Path

import msoffcrypto
import openpyxl

from config import (
    EXCEL_PASSWORD,
    SKIP_SHEETS,
    PINGPONG_KEYWORDS,
    STANDARD_COLS,
    PINGPONG_COLS,
    CURRENCY_MAP,
    ACCOUNT_TYPE_RULES,
)


@dataclass
class Transaction:
    """A single transaction row from a detail sheet."""
    date: str
    sheet_name: str
    account_name: str
    currency: str
    account_type: str
    summary: str
    category: str
    income: float
    expense: float
    balance: float
    is_pingpong: bool
    row_text: str = ""  # Full row text for improved transfer detection


@dataclass
class ParsedSheet:
    """Parsed result from one detail sheet."""
    sheet_name: str
    account_name: str
    currency: str
    account_type: str
    is_pingpong: bool
    transactions: list[Transaction]
    last_balance: float  # Last row's balance (original currency)
    balance_rmb: float  # RMB balance from 日报汇总 (more reliable)


@dataclass
class DailyAccount:
    """Per-account data from 日报汇总 sheet."""
    name: str           # Account name from B column
    prev_balance: float # D column: 昨日余额 (RMB for CNY, local for foreign)
    income: float       # E column: 本日收款
    expense: float      # F column: 本日付款
    balance: float      # G column: 本日余额
    exchange_rate: float = 0.0  # K/L column: exchange rate (0 for CNY)
    source_sheets: list[str] = None  # Sub-sheet names from formulas


def decrypt_excel(filepath: str) -> openpyxl.Workbook:
    """Decrypt an encrypted Excel file and return the workbook."""
    source_path = Path(filepath).expanduser()
    with open(source_path, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=EXCEL_PASSWORD)
        decrypted = io.BytesIO()
        file.decrypt(decrypted)
        decrypted.seek(0)
        wb = openpyxl.load_workbook(decrypted, data_only=True)
    return wb


def is_pingpong_sheet(sheet_name: str) -> bool:
    """Check if a sheet uses Pingpong layout based on its name."""
    name_lower = sheet_name.lower()
    return any(kw.lower() in name_lower for kw in PINGPONG_KEYWORDS)


def extract_currency(sheet_name: str) -> str:
    """Extract currency code from sheet name.

    Checks for currency keywords anywhere in the name (not just suffix).
    This handles variants like "美元户", "人民币CNY", "澳元AUD".
    """
    # First try exact suffix match
    for suffix, code in CURRENCY_MAP.items():
        if sheet_name.endswith(suffix):
            return code
    # Then try keyword-in-name match
    for keyword, code in CURRENCY_MAP.items():
        if keyword in sheet_name:
            return code
    # Default to CNY if no keyword matched
    return "CNY"


def classify_account_type(sheet_name: str) -> str:
    """Classify account type based on sheet name keywords."""
    name_lower = sheet_name.lower()
    for keywords, type_name in ACCOUNT_TYPE_RULES:
        if any(kw.lower() in name_lower for kw in keywords):
            return type_name
    return "其他"


def safe_float(value) -> float:
    """Convert a cell value to float, returning 0.0 for None or invalid values."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def safe_str(value) -> str:
    """Convert a cell value to string, returning empty string for None."""
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(ws) -> tuple[int, bool]:
    """Find the header row by searching for '日期' or '创建时间'.

    Returns (header_row_number, is_pingpong_layout).
    """
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        cell_values = [str(cell.value or "").strip() for cell in row]
        # Pingpong layout: has '创建时间'
        if "创建时间" in cell_values:
            return row[0].row, True
        # Standard layout: has '日期'
        if "日期" in cell_values:
            return row[0].row, False
    return 0, False


def _detect_columns(ws, header_row: int) -> dict[str, int]:
    """Detect column positions from the header row.

    Standard layout: 账户, 日期, 摘要, 类别, 收入, 支出, 余额(USD), ...
    Pingpong layout: 账户, 创建时间, 店铺所在地区, 店铺名称, 备注, 收入, 支出, 余额, ...
    Also accepts variants: 明细=摘要, 分类=类别, 入账=收入, 出账=支出, 余额-*=余额
    Only scans columns up to the one after '支出' to avoid matching '余额（本位币）'.
    """
    cols = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False):
        max_col = 0
        for cell in row:
            val = str(cell.value or "").strip()
            if val in ("支出", "出账"):
                max_col = cell.column + 1  # Scan up to the column after expense
                if "expense" not in cols:
                    cols["expense"] = cell.column
            elif val in ("日期", "创建时间"):
                if "date" not in cols:
                    cols["date"] = cell.column
            elif val in ("摘要", "明细"):
                if "summary" not in cols:
                    cols["summary"] = cell.column
            elif val in ("类别", "分类"):
                if "category" not in cols:
                    cols["category"] = cell.column
            elif val == "备注":
                # Always update to the LAST "备注" column (L col > E col for pingpong)
                cols["remark_last"] = cell.column
                if "remark" not in cols:
                    cols["remark"] = cell.column
            elif val in ("收入", "入账"):
                if "income" not in cols:
                    cols["income"] = cell.column

        # Second pass: find balance within the limited column range
        for cell in row:
            if max_col and cell.column > max_col:
                continue
            val = str(cell.value or "").strip()
            if val == "余额" or val.startswith("余额") or val.startswith("余额("):
                if "balance" not in cols:
                    cols["balance"] = cell.column

    # Balance column: if not found in header, it's the column after expense
    if "balance" not in cols and "expense" in cols:
        cols["balance"] = cols["expense"] + 1

    return cols


def _find_compound_total(ws, balance_col: int) -> float | None:
    """For compound sheets with sub-accounts, find the total balance.

    Two compound patterns exist:
    1. Summary-section type (深圳主体对公美元户, 子公司公账-工商银行):
       - Summary rows at top with sub-account names and balances
       - Detail sections below
    2. Section-per-account type (公账-江苏、建行、招行):
       - Each sub-account has its own section with a 合计 row

    Strategy: Sum ALL 合计 rows throughout the sheet.
    If there are 2+ 合计 rows, this is a compound sheet.
    """
    totals = []
    for row in ws.iter_rows(min_row=1, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        for cell in row:
            val = str(cell.value or "").strip()
            if val == "合计":
                balance = safe_float(cells.get(balance_col))
                totals.append(balance)
                break  # Only one 合计 per row

    if len(totals) >= 2:
        return sum(totals)

    return None


def parse_sheet(ws, sheet_name: str, data_row_start: int = None, data_row_end: int = None) -> ParsedSheet:
    """Parse a single detail sheet into standardized transactions.

    Args:
        data_row_start: Optional minimum row number (absolute) to include.
        data_row_end: Optional maximum row number (absolute) to include.
    """
    is_pp = is_pingpong_sheet(sheet_name)
    currency = extract_currency(sheet_name)
    account_type = classify_account_type(sheet_name)
    account_name = sheet_name

    # Find header row and detect columns dynamically
    header_row, detected_pp = _find_header_row(ws)
    if header_row == 0:
        # Fallback: no header found, use default layout
        cols = PINGPONG_COLS if is_pp else STANDARD_COLS
        data_start = 2
    else:
        is_pp = detected_pp or is_pp
        cols = _detect_columns(ws, header_row)
        data_start = header_row + 1

    # For composite sheets (子公司公账-工商银行, 深圳主体对公美元户),
    # the first header row lacks C=摘要. Sub-account sections below have
    # their own headers with C=摘要. We re-detect columns on every header.
    has_summary = "summary" in cols
    remark_last_col = cols.get("remark_last", 0)

    transactions = []
    last_balance = 0.0

    for row in ws.iter_rows(min_row=data_start, values_only=False):
        row_num = row[0].row
        if data_row_start is not None and row_num < data_row_start:
            continue
        if data_row_end is not None and row_num > data_row_end:
            break
        cells = {cell.column: cell.value for cell in row}

        # Detect sub-account header rows: a row containing literal "摘要" or "明细"
        # as a cell value is a header, not data. Re-detect columns if needed.
        row_text_vals = [str(cell.value or "").strip() for cell in row]
        if "摘要" in row_text_vals or "明细" in row_text_vals:
            if not has_summary or "摘要" == safe_str(cells.get(cols.get("summary", 0))):
                new_cols = _detect_columns(ws, row_num)
                if "summary" in new_cols:
                    cols.update(new_cols)
                    remark_last_col = cols.get("remark_last", 0)
                    has_summary = True
            continue  # Header row is not data

        date_val = cells.get(cols.get("date", 0))
        if date_val is None:
            continue

        income_val = safe_float(cells.get(cols.get("income", 0)))
        expense_val = safe_float(cells.get(cols.get("expense", 0)))
        balance_val = safe_float(cells.get(cols.get("balance", 0)))

        if income_val == 0 and expense_val == 0 and balance_val == 0:
            continue

        date_str = _format_date(date_val)

        summary_col = cols.get("summary", cols.get("remark", 0))
        category_col = cols.get("category", cols.get("remark", 0))

        summary = safe_str(cells.get(summary_col))
        category = safe_str(cells.get(category_col))

        # Pingpong layout: prefer L column (remark_last) for summary if it has data
        if remark_last_col and is_pp:
            l_val = safe_str(cells.get(remark_last_col))
            if l_val:
                summary = l_val

        # Build full row text for comprehensive transfer detection
        row_text_parts = []
        for cell in row:
            val = safe_str(cell.value)
            if val:
                row_text_parts.append(val)
        row_text = " ".join(row_text_parts)

        txn = Transaction(
            date=date_str,
            sheet_name=sheet_name,
            account_name=account_name,
            currency=currency,
            account_type=account_type,
            summary=summary,
            category=category,
            income=income_val,
            expense=expense_val,
            balance=balance_val,
            is_pingpong=is_pp,
            row_text=row_text,
        )
        transactions.append(txn)
        last_balance = balance_val

    # For compound sheets, override last_balance with 合计 total
    balance_col = cols.get("balance", 0)
    if balance_col:
        compound_total = _find_compound_total(ws, balance_col)
        if compound_total is not None:
            last_balance = compound_total

    return ParsedSheet(
        sheet_name=sheet_name,
        account_name=account_name,
        currency=currency,
        account_type=account_type,
        is_pingpong=is_pp,
        transactions=transactions,
        last_balance=last_balance,
        balance_rmb=0.0,  # Overridden by parse_all_sheets from 日报汇总
    )


def _format_date(value) -> str:
    """Format a date value to YYYY-MM-DD string."""
    if value is None:
        return ""

    # If it's already a datetime
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    # If it's a string
    val_str = str(value).strip()
    if not val_str:
        return ""

    # Try to parse common date formats
    # Excel date serial number
    try:
        from datetime import datetime, timedelta
        serial = float(val_str)
        # Excel epoch is 1899-12-30
        base_date = datetime(1899, 12, 30)
        result = base_date + timedelta(days=serial)
        return result.strftime("%Y-%m-%d")
    except (ValueError, TypeError, OverflowError):
        pass

    # Already formatted date string
    date_patterns = [
        (r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", "{0}-{1:0>2}-{2:0>2}"),
        (r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", "{2}-{1:0>2}-{0:0>2}"),
    ]
    for pattern, fmt in date_patterns:
        m = re.match(pattern, val_str)
        if m:
            return fmt.format(*m.groups())

    return val_str


def extract_summary_balances(wb: openpyxl.Workbook) -> dict[str, float]:
    """Extract account balances in RMB from 日报汇总 sheet.

    Returns dict mapping account name patterns to their RMB balance (G column).
    Only reads rows BEFORE the 总计 row to avoid duplicate entries in the
    exchange rate section below it.
    """
    balances: dict[str, float] = {}
    if "日报汇总" not in wb.sheetnames:
        return balances

    ws = wb["日报汇总"]

    # Find the total row to set upper bound
    max_row = 97
    for row in ws.iter_rows(min_row=1, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        a_val = str(cells.get(1, "")).strip()
        if "总计" in a_val:
            max_row = row[0].row
            break

    for row in ws.iter_rows(min_row=5, max_row=max_row - 1, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        name_b = safe_str(cells.get(2))
        name_a = safe_str(cells.get(1))
        name = name_b or name_a
        if not name:
            continue
        if any(skip in name for skip in ("总计", "开户行", "资金汇总", "外币户", "公账户", "公司货币")):
            continue
        g_val = safe_float(cells.get(7))  # G column = 本日余额
        balances[name] = g_val

    return balances


def extract_report_date(wb: openpyxl.Workbook) -> str:
    """Extract the report date from 日报汇总 sheet (Row 2, column K)."""
    if "日报汇总" not in wb.sheetnames:
        return ""
    ws = wb["日报汇总"]
    # J2 = "填报日期:", K2 = actual date value
    date_val = ws.cell(2, 11).value
    return _format_date(date_val)


def _extract_sheet_refs_from_formula(formula: str) -> list[str]:
    """Extract sub-sheet names from SUMIF formulas.

    Handles patterns like:
        SUMIF('SheetName'!B:B, ...)
        SUMIF(SheetName!B:B, ...)
    """
    refs = []
    for m in re.finditer(r"SUMIF\('([^']+)'!", formula):
        refs.append(m.group(1))
    for m in re.finditer(r"SUMIF\((\w+)!", formula):
        refs.append(m.group(1))
    return refs


def _extract_lower_row_ref(formula: str) -> int | None:
    """Extract row number from lower-section reference like =E98*$L$98.

    Returns the row number (e.g., 98) or None.
    """
    m = re.match(r"=E(\d+)", str(formula))
    if m:
        return int(m.group(1))
    m = re.match(r"=F(\d+)", str(formula))
    if m:
        return int(m.group(1))
    return None


def extract_account_sheet_mapping(wb: openpyxl.Workbook) -> dict[str, list[str]]:
    """Extract the mapping from 日报汇总 account names to sub-sheet names via formulas.

    Must be called with a workbook loaded with data_only=False (formulas preserved).

    Returns dict: {日报汇总_account_name: [sub_sheet_names]}
    """
    mapping: dict[str, list[str]] = {}

    if "日报汇总" not in wb.sheetnames:
        return mapping

    ws = wb["日报汇总"]

    # Find the total row
    total_row = None
    for row in ws.iter_rows(min_row=1, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        a_val = str(cells.get(1, "")).strip()
        if "总计" in a_val:
            total_row = row[0].row
            break

    # Build lower-section mapping: row -> sub-sheet names
    lower_sheet_map: dict[int, list[str]] = {}
    if total_row is not None:
        for r in range(total_row + 1, ws.max_row + 1):
            e_formula = str(ws.cell(r, 5).value or "")
            f_formula = str(ws.cell(r, 6).value or "")
            refs = _extract_sheet_refs_from_formula(e_formula) + _extract_sheet_refs_from_formula(f_formula)
            if refs:
                lower_sheet_map[r] = list(dict.fromkeys(refs))

    # Parse upper section: extract account name -> sub-sheet mapping from formulas
    upper_end = (total_row or 97) - 1
    for r in range(5, upper_end + 1):
        col_b = safe_str(ws.cell(r, 2).value)
        col_a = safe_str(ws.cell(r, 1).value)
        name = col_b or col_a
        if not name:
            continue
        if any(skip in name for skip in ("总计", "开户行", "资金汇总", "外币户", "公账户", "公司货币")):
            continue

        e_formula = str(ws.cell(r, 5).value or "")
        f_formula = str(ws.cell(r, 6).value or "")

        # Direct SUMIF reference
        direct_refs = _extract_sheet_refs_from_formula(e_formula) + _extract_sheet_refs_from_formula(f_formula)
        if direct_refs:
            mapping[name] = list(dict.fromkeys(direct_refs))
        else:
            # Lower-section reference (e.g., =E98*$L$98)
            lower_row = _extract_lower_row_ref(e_formula) or _extract_lower_row_ref(f_formula)
            if lower_row and lower_row in lower_sheet_map:
                mapping[name] = lower_sheet_map[lower_row]

    return mapping


def extract_daily_accounts(wb: openpyxl.Workbook) -> tuple[list[DailyAccount], dict[str, float]]:
    """Extract per-account daily data from 日报汇总 sheet (upper RMB section only).

    Returns:
        Tuple of (list of DailyAccount in RMB, exchange rates dict)
    """
    accounts: list[DailyAccount] = []
    rates: dict[str, float] = {"CNY": 1.0}

    if "日报汇总" not in wb.sheetnames:
        return accounts, rates

    ws = wb["日报汇总"]

    # Find the total row (upper section ends here)
    total_row = None
    for row in ws.iter_rows(min_row=1, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        a_val = str(cells.get(1, "")).strip()
        if "总计" in a_val:
            total_row = row[0].row
            break

    # Parse upper section (RMB-converted values, rows 5 to total-1)
    upper_end = (total_row or 97) - 1
    for r in range(5, upper_end + 1):
        col_b = safe_str(ws.cell(r, 2).value)
        col_a = safe_str(ws.cell(r, 1).value)
        name = col_b or col_a
        if not name:
            continue
        if any(skip in name for skip in ("总计", "开户行", "资金汇总", "外币户", "公账户", "公司货币")):
            continue

        prev_bal = safe_float(ws.cell(r, 4).value)   # D: 昨日余额
        income = safe_float(ws.cell(r, 5).value)       # E: 本日收款
        expense = safe_float(ws.cell(r, 6).value)      # F: 本日付款
        cur_bal = safe_float(ws.cell(r, 7).value)       # G: 本日余额

        if abs(prev_bal) < 0.001 and abs(income) < 0.001 and abs(expense) < 0.001 and abs(cur_bal) < 0.001:
            continue

        accounts.append(DailyAccount(
            name=name,
            prev_balance=prev_bal,
            income=income,
            expense=expense,
            balance=cur_bal,
        ))

    # Extract exchange rates from lower section (after total row)
    if total_row is not None:
        for r in range(total_row + 1, ws.max_row + 1):
            col_b = safe_str(ws.cell(r, 2).value)
            col_a = safe_str(ws.cell(r, 1).value)
            name = col_b or col_a
            if not name:
                continue
            rate = safe_float(ws.cell(r, 12).value)   # L column: exchange rate
            if rate > 0 and rate != 1.0:
                for keyword, code in CURRENCY_MAP.items():
                    if keyword in name and code not in ("CNY", "CNH"):
                        if code not in rates:
                            rates[code] = rate

    return accounts, rates


# Mapping from sheet names to 日报汇总 account name patterns.
# Each entry: (sheet_name_substring, summary_name_substring)
# Both substrings must match for a successful mapping.
_ACCOUNT_MATCH_RULES: list[tuple[str, str, bool]] = [
    # (sheet_name_pattern, summary_name_pattern, is_composite)
    # is_composite=True means match multiple summary entries and sum them
    # For non-composite, only the FIRST match is used

    # Citibank multi-currency
    ("Citibank-美元", "Citibank -美元", False),
    ("Citibank-欧元", "Citibank -欧元", False),
    ("Citibank-加元", "Citibank -加元", False),
    ("Citibank-英镑", "Citibank -英镑", False),
    ("Citibank-人民币", "Citibank", False),
    # 德拉姆 pingpong multi-currency (must use 德拉姆 prefix to avoid matching 领头羊)
    ("德拉姆pingpong-美元", "德拉姆（香港）-pingpong-美元", False),
    ("德拉姆pingpong-加元", "德拉姆（香港）-pingpong-加元", False),
    ("德拉姆pingpong-欧元", "德拉姆（香港）-pingpong-欧元", False),
    ("德拉姆pingpong-英镑", "德拉姆（香港）-pingpong-英镑", False),
    ("德拉姆pingpong-墨西哥币", "德拉姆（香港）-pingpong-墨西哥", False),
    ("德拉姆pingpong-澳元", "德拉姆（香港）-pingpong-澳元", False),
    ("德拉姆pingpong日元", "德拉姆（香港）-pingpong-日元", False),
    ("德拉姆pingpong-人民币", "德拉姆（香港）有限公司-pingpong-人民币", False),
    # 领头羊 pingpong
    ("领头羊pingpong-美元", "领头羊-pingpong-美元", False),
    ("领头羊pingpong-欧元", "领头羊-pingpong-欧元", False),
    ("领头羊pingpong-加元", "领头羊-pingpong-加元", False),
    ("领头羊pingpong-英镑", "领头羊-pingpong-英镑", False),
    ("领头羊pingpong-墨西哥", "领头羊-pingpong-墨西哥", False),
    # 汇丰
    ("汇丰-美元", "领头羊-汇丰-美元", False),
    ("汇丰-港币", "领头羊-汇丰-港币", False),
    ("汇丰-人民币", "领头羊-汇丰-人民币", False),
    # 德拉姆寻汇
    ("德拉姆寻汇-人民币", "德拉姆（香港）寻汇-人民币", False),
    # 领头羊中信
    ("领头羊中信FT", "领头羊-中信", False),
    # 德拉姆中信
    ("德拉姆中信FT", "德拉姆香港-中信", False),
    # 光子易虚拟卡
    ("德拉姆-光子易信用卡", "有限公司-光子易虚拟卡", False),
    ("领头羊-光子易信用卡", "领头羊-光子易", False),
    # pingpong虚拟卡
    ("pingpong虚拟信用卡", "pingpong虚拟卡", False),
    # pingpong结汇
    ("pingpong结汇账户(hedging x)-美元", "hedging x)-美元", False),
    # Composite accounts: sum multiple summary entries
    ("pingpong沃尔玛收款", "沃尔玛收款", True),
    ("pingpong-TK收款", "TK收款", True),
    ("钱海账户", "钱海", False),
    ("paypal账户", "PayPal", True),
    ("P卡账户", "P卡", True),
    # 现金 (compound: 人民币 + 日元)
    ("现金", "现金", True),
]


def _match_summary_balance(
    sheet_name: str, summary_balances: dict[str, float]
) -> float | None:
    """Find the matching RMB balance(s) from 日报汇总 for a given sheet name.

    A sheet may match multiple summary accounts (e.g., "pingpong沃尔玛收款"
    matches 5 individual Walmart sub-accounts). Returns the sum of all matches.
    Uses only the FIRST matching rule to avoid cross-rule accumulation.
    """
    # Direct match first
    if sheet_name in summary_balances:
        return summary_balances[sheet_name]

    # Find the first matching rule and collect matching summary accounts
    for sheet_pattern, summary_pattern, is_composite in _ACCOUNT_MATCH_RULES:
        if sheet_pattern in sheet_name:
            matches: list[float] = []
            for summary_name, balance in summary_balances.items():
                if summary_pattern in summary_name:
                    matches.append(balance)
                    if not is_composite:
                        break  # Only take first match for non-composite
            if matches:
                return sum(matches)

    # Fallback: check if sheet name is substring of any summary name
    for summary_name, balance in summary_balances.items():
        if sheet_name in summary_name or summary_name in sheet_name:
            return balance

    return None


def extract_exchange_rates(wb: openpyxl.Workbook) -> dict[str, float]:
    """Extract exchange rates from 日报汇总 sheet.

    Exchange rates are in the L column of the section AFTER the total row (row 97).
    Each row has the account name in B column and the rate in L column.
    """
    rates = {}
    if "日报汇总" not in wb.sheetnames:
        return rates

    ws = wb["日报汇总"]

    # Find the total row first
    total_row = None
    for row in ws.iter_rows(min_row=1, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        a_val = str(cells.get(1, "")).strip()
        if "总计" in a_val:
            total_row = row[0].row
            break

    if total_row is None:
        total_row = 97  # fallback

    # Read exchange rates from rows after the total
    for row in ws.iter_rows(min_row=total_row + 1, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        rate_val = safe_float(cells.get(12))  # L column
        if rate_val and rate_val > 0 and rate_val != 1.0:
            # Try to identify currency from B column account name
            account_name = safe_str(cells.get(2))
            if not account_name:
                account_name = safe_str(cells.get(1))
            for keyword, code in CURRENCY_MAP.items():
                if keyword in account_name and code not in ("CNY", "CNH"):
                    if code not in rates:
                        rates[code] = rate_val

    return rates


def parse_all_sheets(filepath: str) -> tuple[list[ParsedSheet], dict[str, float]]:
    """Parse all detail sheets from the encrypted Excel file.

    Returns:
        Tuple of (list of parsed sheets, exchange rates dict)
    """
    wb = decrypt_excel(filepath)
    rates = extract_exchange_rates(wb)
    summary_balances = extract_summary_balances(wb)

    parsed = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        result = parse_sheet(ws, sheet_name)
        if result.transactions:
            # Use 日报汇总 balance as authoritative RMB value
            rmb_balance = _match_summary_balance(sheet_name, summary_balances)
            result = ParsedSheet(
                sheet_name=result.sheet_name,
                account_name=result.account_name,
                currency=result.currency,
                account_type=result.account_type,
                is_pingpong=result.is_pingpong,
                transactions=result.transactions,
                last_balance=result.last_balance,
                balance_rmb=rmb_balance if rmb_balance is not None else result.last_balance * rates.get(result.currency, 1.0),
            )
            parsed.append(result)

    wb.close()
    return parsed, rates
