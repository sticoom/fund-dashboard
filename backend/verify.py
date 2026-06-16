"""Formula-aware reconciliation engine.

Parses 日报汇总 formulas to trace exact data flow:
  Lower section (rows 98+): SUMIF referencing sub-sheet columns → original currency values
  Upper section (rows 5-96): =E{lower}*$L${lower} → RMB-converted values
  Exchange rate in column L of lower section determines actual currency.

Verification chain:
  sub-sheet transactions → SUMIF result (lower section) → × exchange rate → RMB (upper section)
"""

import io
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import msoffcrypto
import openpyxl

from parser import (
    extract_report_date,
    parse_sheet,
    safe_float,
    safe_str,
)
from classifier import is_transfer, TRANSFER_CATEGORIES as TRANSFER_KEYWORDS, is_suspected_transfer, find_cross_sheet_pairs
from config import EXCEL_PASSWORD, SKIP_SHEETS, CURRENCY_MAP

TOLERANCE = 0.5


def _sumif_criteria_range(formula_str: str) -> str:
    """Extract the criteria range string from a SUMIF formula.

    Returns the first argument (criteria range) of =SUMIF(range, criteria, sum_range).
    """
    formula = str(formula_str or "")
    if not formula.startswith("=SUMIF"):
        return ""
    inner = formula[len("=SUMIF("):]
    if inner.endswith(")"):
        inner = inner[:-1]
    depth = 0
    in_quote = False
    for i, ch in enumerate(inner):
        if ch == "'":
            in_quote = not in_quote
        elif ch == "(" and not in_quote:
            depth += 1
        elif ch == ")" and not in_quote:
            depth -= 1
        elif ch == "," and depth == 0 and not in_quote:
            return inner[:i].strip()
    return ""


def _get_sumif_row_range(formula_str: str) -> tuple[int | None, int | None]:
    """Extract row range from a SUMIF formula's criteria range.

    Returns (start_row, end_row) or (None, None) for full-column references.
    """
    crit_range = _sumif_criteria_range(formula_str)
    if not crit_range:
        return None, None
    _, start, end, _ = _parse_sheet_range(crit_range)
    return start, end


def _eval_sumif(wb: openpyxl.Workbook, formula: str, date: str) -> tuple[float, float]:
    """Evaluate a SUMIF formula directly by reading the detail sheet.

    Returns (total, transfer_amount) for rows matching the given date.
    Transfer detection checks all text columns for keywords like '往来'.
    """
    if not formula or not str(formula).startswith("=SUMIF"):
        return 0.0, 0.0
    formula = str(formula)

    # Extract SUMIF arguments: (range1, criteria, range2)
    # Remove =SUMIF( prefix and trailing )
    inner = formula[len("=SUMIF("):]
    if inner.endswith(")"):
        inner = inner[:-1]

    # Split on commas, but respect quoted sheet names with commas inside
    # e.g., SUMIF('Sheet, Inc.'!B:B, K2, 'Sheet, Inc.'!E:E)
    parts = []
    depth = 0
    current = ""
    in_quote = False
    for ch in inner:
        if ch == "'" and not in_quote:
            in_quote = True
            current += ch
        elif ch == "'" and in_quote:
            in_quote = False
            current += ch
        elif ch == "(" and not in_quote:
            depth += 1
            current += ch
        elif ch == ")" and not in_quote:
            depth -= 1
            current += ch
        elif ch == "," and depth == 0 and not in_quote:
            parts.append(current.strip())
            current = ""
        else:
            current += ch
    if current.strip():
        parts.append(current.strip())

    if len(parts) < 3:
        return 0.0, 0.0

    criteria_range = parts[0]  # e.g. 'SheetName'!B3:B19
    # criteria = parts[1]  # e.g. $K$2 -- we already know it's the date
    sum_range = parts[2]  # e.g. 'SheetName'!E3:E19

    # Parse sheet name and range from criteria_range
    sheet_name, crit_start_row, crit_end_row, crit_col = _parse_sheet_range(criteria_range)
    if not sheet_name:
        return 0.0, 0.0

    # Parse sheet name and range from sum_range
    _, sum_start_row, sum_end_row, sum_col = _parse_sheet_range(sum_range)

    if sheet_name not in wb.sheetnames:
        return 0.0, 0.0

    ws = wb[sheet_name]

    # Determine row range
    start_row = max(crit_start_row or 1, sum_start_row or 1)
    end_row = max(crit_end_row or ws.max_row, sum_end_row or ws.max_row)

    total = 0.0
    transfer_total = 0.0
    for r in range(start_row, end_row + 1):
        # Check date match in criteria column
        cell_val = ws.cell(r, crit_col).value
        if cell_val is None:
            continue
        cell_date = _format_date_val(cell_val)
        if cell_date == date:
            val = safe_float(ws.cell(r, sum_col).value)
            total += val
            # Transfer detection: check all text columns for '往来'
            row_text = " ".join(
                str(ws.cell(r, c).value or "")
                for c in range(1, ws.max_column + 1)
            )
            if any(kw in row_text for kw in TRANSFER_KEYWORDS):
                transfer_total += val

    return total, transfer_total


def _parse_sheet_range(range_str: str) -> tuple[str, int | None, int | None, int]:
    """Parse a range reference like 'SheetName'!B3:B19 or SheetName!$B:$B.

    Returns (sheet_name, start_row_or_None, end_row_or_None, col_index).
    """
    sheet_name = ""
    range_part = range_str

    if "'" in range_str:
        m = re.match(r"'([^']+)'" + r"!(.*)", range_str)
        if m:
            sheet_name = m.group(1)
            range_part = m.group(2)
    elif "!" in range_str:
        idx = range_str.index("!")
        sheet_name = range_str[:idx]
        range_part = range_str[idx + 1:]

    # Remove $ signs
    range_part = range_part.replace("$", "")

    # Parse column range: B3:B19 or B:B
    m = re.match(r"([A-Z]+)(\d*):([A-Z]+)(\d*)", range_part, re.IGNORECASE)
    if not m:
        return sheet_name, None, None, 0

    col_letter = m.group(1).upper()
    start_row = int(m.group(2)) if m.group(2) else None
    end_row = int(m.group(4)) if m.group(4) else None

    col_idx = _col_letter_to_idx(col_letter)

    return sheet_name, start_row, end_row, col_idx


def _format_date_val(value) -> str:
    """Format a cell value as YYYY-MM-DD date string."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    val_str = str(value).strip()
    if not val_str:
        return ""
    # Try numeric (Excel serial)
    try:
        from datetime import datetime, timedelta
        serial = float(val_str)
        base_date = datetime(1899, 12, 30)
        return (base_date + timedelta(days=serial)).strftime("%Y-%m-%d")
    except (ValueError, TypeError, OverflowError):
        pass
    # String patterns
    for pattern, fmt in [
        (r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", "{0}-{1:0>2}-{2:0>2}"),
        (r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", "{2}-{1:0>2}-{0:0>2}"),
    ]:
        m = re.match(pattern, val_str)
        if m:
            return fmt.format(*m.groups())
    return val_str


def _decrypt(filepath: str, data_only: bool) -> openpyxl.Workbook:
    source_path = Path(filepath).expanduser()
    with open(source_path, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=EXCEL_PASSWORD)
        buf = io.BytesIO()
        file.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=data_only)


def _parse_sumif(formula: str):
    """Extract sheet_name, date_col_range, sum_col_range from SUMIF formula.

    e.g. =SUMIF('德拉姆pingpong-加元'!$B:$B,$K$2,'德拉姆pingpong-加元'!F:F)
    returns ('德拉姆pingpong-加元', 'B:B', 'F:F')
    """
    if not formula or not str(formula).startswith("=SUMIF"):
        return None
    formula = str(formula)
    # Match SUMIF(sheet!col_range, criteria, sheet!col_range) or direct cell ref
    m = re.match(
        r"=SUMIF\('([^']+)'!\$(\w+):\$(\w+),.*?'[^']+'!(\w+):(\w+)\)",
        formula,
    )
    if m:
        sheet = m.group(1)
        sum_col = m.group(4)
        return sheet, sum_col
    m = re.match(r"=SUMIF\('([^']+)'!\$(\w+):\$(\w+),.*?[^']+'!(\w+):(\w+)\)", formula)
    if m:
        return m.group(1), m.group(3)
    m = re.match(r"=SUMIF\((\w+)!\$(\w+):\$(\w+),.*?\1!(\w+):(\w+)\)", formula)
    if m:
        return m.group(1), m.group(4)
    m = re.match(r"=SUMIF\('([^']+)'!(\w+):(\w+),.*?'[^']+'!(\w+):(\w+)\)", formula)
    if m:
        return m.group(1), m.group(4)
    # range variant e.g. pingpong沃尔玛收款!B3:B19
    m = re.match(r"=SUMIF\('([^']+)'!(\w+\d+:\w+\d+),.*?'[^']+'!(\w+\d+:\w+\d+)\)", formula)
    if m:
        return m.group(1), m.group(3).split(":")[0][:1]
    m = re.match(r"=SUMIF\((\w+)!(\w+):(\w+),.*?\1!(\w+):(\w+)\)", formula)
    if m:
        return m.group(1), m.group(4)
    return None


def _parse_direct_ref(formula: str):
    """Parse direct cell reference like =深圳主体对公美元户!I6 or ='子公司公账-工商银行'!I18."""
    if not formula or not str(formula).startswith("="):
        return None
    formula = str(formula)
    m = re.match(r"='([^']+)'!(\w+)(\d+)", formula)
    if m:
        return m.group(1), m.group(2), int(m.group(3))
    m = re.match(r"=(\w+)!(\w+)(\d+)", formula)
    if m:
        return m.group(1), m.group(2), int(m.group(3))
    return None


def _col_letter_to_idx(letter: str) -> int:
    """Convert column letter to 1-based index: A=1, B=2, ..., Z=26, AA=27."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _extract_filename_date(filepath: str) -> str | None:
    """Extract date from filename like '2026年5月资金流动表5.13.xlsx' → '2026-05-13'."""
    name = Path(filepath).stem
    # Match: "2026年5月资金流动表5.13" → year=2026, month=5, day=13
    m = re.search(r'(\d{4})年(\d+)月.*?表(\d+)\.(\d+)', name)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(4)):02d}"
    # Fallback: "5.13" after "表" means month.day
    m = re.search(r'表(\d+)\.(\d+)', name)
    if m:
        ym = re.search(r'(\d{4})年(\d+)月', name)
        if ym:
            return f"{ym.group(1)}-{int(ym.group(2)):02d}-{int(m.group(2)):02d}"
        from datetime import datetime
        return f"{datetime.now().year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def verify(filepath: str, original_filename: str | None = None) -> dict:
    """Run formula-aware reconciliation.

    Args:
        filepath: Path to the encrypted Excel file.
        original_filename: Original upload filename (for date extraction).
    """
    wb_data = _decrypt(filepath, data_only=True)
    wb_formulas = _decrypt(filepath, data_only=False)

    # Date: K2 (sheet_date) is authoritative, filename_date is cross-check
    fname = original_filename or filepath
    filename_date = _extract_filename_date(fname)
    sheet_date = extract_report_date(wb_data)
    date = sheet_date or filename_date
    if not date:
        from datetime import datetime
        date = datetime.now().strftime("%Y-%m-%d")
    date_mismatch = (filename_date and sheet_date and filename_date != sheet_date)

    ws_data = wb_data["日报汇总"]
    ws_formulas = wb_formulas["日报汇总"]

    # Find total row (公司货币资金 总计)
    total_row = None
    for row in ws_data.iter_rows(min_row=1, max_row=120, values_only=False):
        a_val = str(row[0].value or "").strip()
        if "总计" in a_val:
            total_row = row[0].row
            break
    if total_row is None:
        total_row = 97

    # Read authoritative totals from 公司货币资金 总计 row
    total_prev = safe_float(ws_data.cell(total_row, 4).value)   # D: 昨日余额 = 期初余额
    total_income = safe_float(ws_data.cell(total_row, 5).value)  # E: 本日收款
    total_expense = safe_float(ws_data.cell(total_row, 6).value) # F: 本日付款
    total_balance = safe_float(ws_data.cell(total_row, 7).value) # G: 本日余额

    # ---- Parse lower section (rows total_row+1 to max_row) ----
    # Each row has: name (B), prev_balance (D), income_formula (E), expense_formula (F), rate (L)
    # Upper section rows with =D{lower}*$L${lower} reference these rows
    lower_rows = {}  # row_num -> {name, sheet, income_col, expense_col, rate, prev_local, income_local, expense_local}
    sheet_to_lower = {}  # sub_sheet_name -> lower_row_info

    for r in range(total_row + 1, ws_data.max_row + 1):
        name = safe_str(ws_data.cell(r, 2).value) or safe_str(ws_data.cell(r, 1).value)
        if not name:
            continue
        rate = safe_float(ws_data.cell(r, 12).value)  # L column
        if rate == 0:
            rate = 1.0
        prev_local = safe_float(ws_data.cell(r, 4).value)  # D column (local currency)
        income_local_cache = safe_float(ws_data.cell(r, 5).value)  # E column (SUMIF cache)
        expense_local_cache = safe_float(ws_data.cell(r, 6).value)  # F column (SUMIF cache)

        # Parse formulas to get sub-sheet and column references
        income_formula = str(ws_formulas.cell(r, 5).value or "")
        expense_formula = str(ws_formulas.cell(r, 6).value or "")

        income_ref = _parse_sumif(income_formula)
        expense_ref = _parse_sumif(expense_formula)

        # Evaluate SUMIF directly from detail sheets (bypasses stale cache)
        if income_formula.startswith("=SUMIF"):
            income_local, income_transfer = _eval_sumif(wb_data, income_formula, date)
        else:
            income_local, income_transfer = income_local_cache, 0.0
        if expense_formula.startswith("=SUMIF"):
            expense_local, expense_transfer = _eval_sumif(wb_data, expense_formula, date)
        else:
            expense_local, expense_transfer = expense_local_cache, 0.0

        # Expense column preserves sign (negative = refund/credit)

        info = {
            "name": name,
            "row": r,
            "rate": rate,
            "prev_local": prev_local,
            "income_local": income_local,
            "expense_local": expense_local,
            "income_transfer": income_transfer,
            "expense_transfer": expense_transfer,
            "income_local_cache": income_local_cache,
            "expense_local_cache": expense_local_cache,
            "income_sheet": income_ref[0] if income_ref else None,
            "income_col": income_ref[1] if income_ref else None,
            "expense_sheet": expense_ref[0] if expense_ref else None,
            "expense_col": expense_ref[1] if expense_ref else None,
            "is_cny": rate == 1.0,
            "income_range": _get_sumif_row_range(income_formula),
            "expense_range": _get_sumif_row_range(expense_formula),
        }

        # Also check for direct cell references
        if not income_ref:
            direct = _parse_direct_ref(income_formula)
            if direct:
                info["income_sheet"] = direct[0]
                info["income_col"] = direct[1]
                info["income_row"] = direct[2]
                info["income_is_direct"] = True
        if not expense_ref:
            direct = _parse_direct_ref(expense_formula)
            if direct:
                info["expense_sheet"] = direct[0]
                info["expense_col"] = direct[1]
                info["expense_row"] = direct[2]
                info["expense_is_direct"] = True

        lower_rows[r] = info
        if info["income_sheet"]:
            sheet_to_lower[info["income_sheet"]] = info

    # ---- Parse upper section (rows 5 to total_row-1) ----
    # Upper rows reference lower rows via =D{lower}*$L${lower}
    upper_rows = []
    for r in range(5, total_row):
        name = safe_str(ws_data.cell(r, 2).value) or safe_str(ws_data.cell(r, 1).value)
        if not name:
            continue
        if any(skip in name for skip in ("开户行", "资金汇总")):
            continue

        prev_rmb = safe_float(ws_data.cell(r, 4).value)
        income_rmb = safe_float(ws_data.cell(r, 5).value)
        expense_rmb = safe_float(ws_data.cell(r, 6).value)
        balance_rmb = safe_float(ws_data.cell(r, 7).value)

        # Parse formulas to find lower-section row reference
        d_formula = str(ws_formulas.cell(r, 4).value or "")
        e_formula = str(ws_formulas.cell(r, 5).value or "")
        f_formula = str(ws_formulas.cell(r, 6).value or "")

        # Check if it references a lower row (foreign currency account)
        lower_ref = None
        m = re.match(r"=D(\d+)\*\$", d_formula)
        if m:
            lower_ref = int(m.group(1))
        if not lower_ref:
            m = re.match(r"=D(\d+)\*", d_formula)
            if m:
                lower_ref = int(m.group(1))

        # Check for direct SUMIF (CNY accounts in upper section)
        direct_sumif = _parse_sumif(e_formula) if e_formula.startswith("=SUMIF") else None
        direct_ref = _parse_direct_ref(e_formula) if not direct_sumif and e_formula.startswith("=") else None

        upper_rows.append({
            "row": r,
            "name": name,
            "prev_rmb": prev_rmb,
            "income_rmb": income_rmb,
            "expense_rmb": expense_rmb,
            "balance_rmb": balance_rmb,
            "lower_ref": lower_ref,
            "direct_sumif": direct_sumif,
            "direct_ref": direct_ref,
        })

    # ---- Now compute verification for each active account ----
    active_sheets = []

    for upper in upper_rows:
        # Skip inactive accounts, but always include composite sub-accounts
        # (those sharing a physical sheet with specific SUMIF row ranges)
        if abs(upper["income_rmb"]) < 0.01 and abs(upper["expense_rmb"]) < 0.01:
            is_composite_sub = False
            if upper["lower_ref"] and upper["lower_ref"] in lower_rows:
                lr = lower_rows[upper["lower_ref"]]
                inc_range = lr.get("income_range", (None, None))
                if inc_range[0] is not None:
                    is_composite_sub = True
            if not is_composite_sub:
                continue

        lower = None
        sheet_name = None
        rate = 1.0
        income_local_reported = 0.0
        expense_local_reported = 0.0
        prev_local_reported = 0.0

        if upper["lower_ref"] and upper["lower_ref"] in lower_rows:
            # Foreign currency account - references lower section
            lower = lower_rows[upper["lower_ref"]]
            sheet_name = lower["income_sheet"] or lower["expense_sheet"] or upper["name"]
            rate = lower["rate"]
            income_local_reported = lower["income_local"]
            expense_local_reported = lower["expense_local"]
            prev_local_reported = lower["prev_local"]
        elif upper["direct_sumif"]:
            sheet_name = upper["direct_sumif"][0]
            income_local_reported = upper["income_rmb"]
            expense_local_reported = upper["expense_rmb"]
            prev_local_reported = upper["prev_rmb"]
        elif upper["direct_ref"]:
            sheet_name = upper["direct_ref"][0]
            income_local_reported = upper["income_rmb"]
            expense_local_reported = upper["expense_rmb"]
            prev_local_reported = upper["prev_rmb"]
        else:
            sheet_name = upper["name"]

        if not sheet_name:
            continue

        # Determine currency from exchange rate
        currency = "CNY"
        if lower and not lower["is_cny"]:
            for kw, code in CURRENCY_MAP.items():
                if kw in (lower["name"] or ""):
                    currency = code
                    break
            else:
                currency = "USD" if rate > 5 else "OTH"

        # Compute from sub-sheet transactions using SUMIF evaluation
        calc_local_income = 0.0
        calc_local_expense = 0.0
        real_income = 0.0
        real_expense = 0.0
        transfer_income = 0.0
        transfer_expense = 0.0
        txns = []

        # Get the actual SUMIF formulas for this upper row
        e_formula_upper = str(ws_formulas.cell(upper["row"], 5).value or "")
        f_formula_upper = str(ws_formulas.cell(upper["row"], 6).value or "")

        if upper["lower_ref"] and upper["lower_ref"] in lower_rows:
            # Foreign currency: use lower section _eval_sumif result
            lower = lower_rows[upper["lower_ref"]]
            calc_local_income = lower["income_local"]
            calc_local_expense = lower["expense_local"]
            transfer_local_income = lower["income_transfer"]
            transfer_local_expense = lower["expense_transfer"]
        elif e_formula_upper.startswith("=SUMIF"):
            # CNY account with direct SUMIF in upper section
            calc_local_income, transfer_local_income = _eval_sumif(wb_data, e_formula_upper, date)
            if f_formula_upper.startswith("=SUMIF"):
                calc_local_expense, transfer_local_expense = _eval_sumif(wb_data, f_formula_upper, date)
            else:
                calc_local_expense, transfer_local_expense = 0.0, 0.0
            # Expense preserves sign (negative values = refunds/credits)
        elif upper["direct_ref"]:
            # Direct cell reference: use formula cache value
            calc_local_income = upper["income_rmb"]
            calc_local_expense = upper["expense_rmb"]
            transfer_local_income = 0.0
            transfer_local_expense = 0.0

        # Parse transactions for display and transfer classification
        if sheet_name in wb_data.sheetnames and sheet_name not in SKIP_SHEETS:
            ws = wb_data[sheet_name]
            # For composite accounts with specific row ranges, filter transactions
            row_start = None
            row_end = None
            if lower:
                inc_range = lower.get("income_range")
                if inc_range and inc_range[0] is not None:
                    row_start, row_end = inc_range
            result = parse_sheet(ws, sheet_name, data_row_start=row_start, data_row_end=row_end)
            for txn in result.transactions:
                if txn.date and txn.date != date:
                    continue
                txns.append({
                    "summary": txn.summary,
                    "category": txn.category,
                    "income": txn.income,
                    "expense": txn.expense,
                    "balance": txn.balance,
                    "is_transfer": is_transfer(txn),
                })

        # For accounts where formulas don't support transfer detection
        # (direct_ref in upper section, or non-SUMIF references in lower section),
        # compute transfers from transaction list instead
        if transfer_local_income == 0 and transfer_local_expense == 0:
            txn_transfer_income = sum(t["income"] for t in txns if t.get("is_transfer"))
            txn_transfer_expense = sum(t["expense"] for t in txns if t.get("is_transfer"))
            if txn_transfer_income > 0 or txn_transfer_expense > 0:
                transfer_local_income = txn_transfer_income
                transfer_local_expense = txn_transfer_expense

        # Real income/expense (excluding transfers) from _eval_sumif
        real_local_income = calc_local_income - transfer_local_income
        real_local_expense = calc_local_expense - transfer_local_expense

        # Verification
        income_ok = abs(calc_local_income - income_local_reported) < TOLERANCE
        expense_ok = abs(calc_local_expense - expense_local_reported) < TOLERANCE

        # Balance check: prev + income - expense = balance
        expected_balance = upper["prev_rmb"] + upper["income_rmb"] - upper["expense_rmb"]
        balance_ok = abs(upper["balance_rmb"] - expected_balance) < TOLERANCE

        calc_rmb_income = calc_local_income * rate
        calc_rmb_expense = calc_local_expense * rate
        real_rmb_income = real_local_income * rate
        real_rmb_expense = real_local_expense * rate
        transfer_rmb_income = transfer_local_income * rate
        transfer_rmb_expense = transfer_local_expense * rate

        real_income_rmb = real_income * rate
        real_expense_rmb = real_expense * rate
        transfer_income_rmb = transfer_income * rate
        transfer_expense_rmb = transfer_expense * rate

        active_sheets.append({
            "sheet_name": sheet_name,
            "summary_name": upper["name"],
            "currency": currency,
            "exchange_rate": rate,
            "prev_local": prev_local_reported,
            "local_income": calc_local_income,
            "local_expense": calc_local_expense,
            "total_rmb_income": calc_rmb_income,
            "total_rmb_expense": calc_rmb_expense,
            "rmb_income": real_rmb_income,
            "rmb_expense": real_rmb_expense,
            "reported_income": upper["income_rmb"],
            "reported_expense": upper["expense_rmb"],
            "reported_balance": upper["balance_rmb"],
            "reported_prev": upper["prev_rmb"],
            "income_ok": income_ok,
            "expense_ok": expense_ok,
            "balance_ok": balance_ok,
            "all_ok": income_ok and expense_ok and balance_ok,
            "income_diff": calc_local_income - income_local_reported,
            "expense_diff": calc_local_expense - expense_local_reported,
            "real_income": real_local_income,
            "real_expense": real_local_expense,
            "real_income_rmb": real_rmb_income,
            "real_expense_rmb": real_rmb_expense,
            "transfer_income": transfer_local_income,
            "transfer_expense": transfer_local_expense,
            "transfer_income_rmb": transfer_rmb_income,
            "transfer_expense_rmb": transfer_rmb_expense,
            "transactions": txns,
            "formula_source": "lower_ref" if lower else ("direct_sumif" if upper["direct_sumif"] else "direct_ref" if upper["direct_ref"] else "unknown"),
        })

    wb_data.close()
    wb_formulas.close()

    # Use computed values for summary (bypasses stale formula cache)
    # Total income/expense (INCLUDING transfers) - matches Excel table
    total_all_income = sum(s["total_rmb_income"] for s in active_sheets)
    total_all_expense = sum(s["total_rmb_expense"] for s in active_sheets)
    # Real income/expense (EXCLUDING transfers) - actual business flow
    total_real_income = sum(s["rmb_income"] for s in active_sheets)
    total_real_expense = sum(s["rmb_expense"] for s in active_sheets)
    # Transfer amounts
    total_transfer_income = sum(s["transfer_income"] * s["exchange_rate"] for s in active_sheets)
    total_transfer_expense = sum(s["transfer_expense"] * s["exchange_rate"] for s in active_sheets)
    # Balance uses total (including transfers)
    computed_balance = total_prev + total_all_income - total_all_expense
    issues = [s for s in active_sheets if not s["all_ok"]]

    # ---- Cross-sheet transfer reconciliation (Layer 2) ----
    # Collect all transactions for cross-sheet pairing
    all_txns = []
    for s in active_sheets:
        for t in s.get("transactions", []):
            all_txns.append(t)

    warnings = []

    # Find suspected transfers (Layer 1.5: keyword-based suspicion)
    for t in all_txns:
        if not t.get("is_transfer") and _txn_is_suspected(t):
            warnings.append({
                "type": "suspected_transfer",
                "sheet": t.get("sheet_name", ""),
                "category": t.get("category", ""),
                "summary": t.get("summary", ""),
                "income": t.get("income", 0),
                "expense": t.get("expense", 0),
                "message": f"[{t.get('category', '')}] {t.get('summary', '')} 疑似内部往来但未标记为往来",
            })

    # Find cross-sheet pairs by amount matching (Layer 2)
    amount_warnings = _find_amount_pairs(active_sheets)
    warnings.extend(amount_warnings)

    # ---- Transfer summary for transfer reconciliation page ----
    transfer_sheets = []
    for s in active_sheets:
        transfer_txns = [t for t in s.get("transactions", []) if t.get("is_transfer")]
        if transfer_txns or s["transfer_income"] > 0.01 or s["transfer_expense"] > 0.01:
            transfer_sheets.append({
                "sheet_name": s["sheet_name"],
                "currency": s["currency"],
                "exchange_rate": s["exchange_rate"],
                "transfer_income": s["transfer_income"],
                "transfer_expense": s["transfer_expense"],
                "transfer_income_rmb": s["transfer_income_rmb"],
                "transfer_expense_rmb": s["transfer_expense_rmb"],
                "net": s["transfer_income"] - s["transfer_expense"],
                "net_rmb": s["transfer_income_rmb"] - s["transfer_expense_rmb"],
                "transactions": transfer_txns,
            })

    total_transfer_income_rmb = sum(s["transfer_income_rmb"] for s in active_sheets)
    total_transfer_expense_rmb = sum(s["transfer_expense_rmb"] for s in active_sheets)
    transfer_diff_rmb = total_transfer_income_rmb - total_transfer_expense_rmb

    total_real_income_rmb = sum(s["real_income_rmb"] for s in active_sheets)
    total_real_expense_rmb = sum(s["real_expense_rmb"] for s in active_sheets)

    transfer_summary = {
        "total_income_rmb": total_transfer_income_rmb,
        "total_expense_rmb": total_transfer_expense_rmb,
        "diff_rmb": transfer_diff_rmb,
        "balanced": abs(transfer_diff_rmb) < 1.0,
        "sheets": transfer_sheets,
    }

    # ---- FX loss analysis ----
    fx_loss = _build_fx_loss_pairs(active_sheets)

    # ---- Three-category view (lightweight packaging of fx_loss for UI) ----
    # balanced:     pairs with no FX loss
    # fx_loss:      pairs with cross-currency FX loss (explainable business cost)
    # unmatched:    transfer txns that found no counterpart in any sheet (data gap)
    fx_pairs = fx_loss["pairs"]
    fx_unmatched = fx_loss.get("unmatched", [])

    balanced_pairs = [p for p in fx_pairs if abs(p["loss"]) < 0.01]
    fx_loss_pairs = [p for p in fx_pairs if abs(p["loss"]) >= 0.01]

    unmatched_net_rmb = round(
        sum(
            u["rmb"] * (1 if u["direction"] == "income" else -1)
            for u in fx_unmatched
        ),
        2,
    )
    # diff_rmb should equal fx_loss + unmatched_net if everything is explained
    explained = abs(transfer_diff_rmb - fx_loss["total_loss"] - unmatched_net_rmb) < 1.0

    transfer_summary["categorized"] = {
        "balanced_pairs": balanced_pairs,
        "fx_loss_pairs": fx_loss_pairs,
        "unmatched": fx_unmatched,
        "unmatched_net_rmb": unmatched_net_rmb,
        "explained": explained,
    }

    return {
        "date": date,
        "filename": Path(filepath).name,
        "filename_date": filename_date,
        "sheet_date": sheet_date,
        "date_mismatch": date_mismatch,
        "summary": {
            "prev_balance": total_prev,
            "balance": total_balance,
            # Total (including transfers) - matches Excel verification
            "total_income": total_all_income,
            "total_expense": total_all_expense,
            # Reported values from Excel (authoritative)
            "income": total_income,
            "expense": total_expense,
            "net_flow": total_income - total_expense,
            # Real (excluding transfers) - actual business flow (RMB)
            "real_income": total_real_income_rmb,
            "real_expense": total_real_expense_rmb,
            "real_net": total_real_income_rmb - total_real_expense_rmb,
            # Transfer detail (RMB)
            "transfer_income": total_transfer_income_rmb,
            "transfer_expense": total_transfer_expense_rmb,
            # Formula cache comparison
            "reported_income": total_income,
            "reported_expense": total_expense,
            "reported_balance": total_balance,
            "expected_balance": total_prev + total_income - total_expense,
            "income_match": abs(total_all_income - total_income) < TOLERANCE,
            "expense_match": abs(total_all_expense - total_expense) < TOLERANCE,
            "balance_match": abs(total_balance - (total_prev + total_income - total_expense)) < TOLERANCE,
        },
        "active_accounts": len(active_sheets),
        "issues_count": len(issues),
        "warnings": warnings,
        "transfer_summary": transfer_summary,
        "fx_loss": fx_loss,
        "sheets": active_sheets,
    }


# ---- Helper functions for cross-sheet reconciliation ----

_SUSPECTED_CATEGORIES = {"投资收益", "投资款"}
_TRANSFER_SUMMARY_PATTERNS = ["转入", "转出", "提回", "划转", "调拨", "提现到"]


def _txn_is_suspected(t: dict) -> bool:
    """Check if a transaction dict is suspected to be an internal transfer."""
    category = (t.get("category") or "").strip()
    summary = t.get("summary") or ""
    income = t.get("income", 0)
    expense = t.get("expense", 0)

    if category in _SUSPECTED_CATEGORIES:
        return True

    if any(kw in summary for kw in _TRANSFER_SUMMARY_PATTERNS):
        if income > 100 or expense > 100:
            return True

    return False


def _find_amount_pairs(active_sheets: list[dict]) -> list[dict]:
    """Find matching expense-income pairs across sheets by amount.

    If neither side of a pair is marked as transfer, flag it as a warning.
    """
    # Collect expenses and incomes with sheet context
    expenses = []  # (amount, sheet_name, txn_dict)
    incomes = []   # (amount, sheet_name, txn_dict)

    for s in active_sheets:
        for t in s.get("transactions", []):
            income = t.get("income", 0)
            expense = t.get("expense", 0)
            if expense > TOLERANCE:
                expenses.append((round(expense, 2), s["sheet_name"], t))
            if income > TOLERANCE:
                incomes.append((round(income, 2), s["sheet_name"], t))

    used_income = set()
    warnings = []

    for e_amt, e_sheet, e_txn in expenses:
        for i_idx, (i_amt, i_sheet, i_txn) in enumerate(incomes):
            if i_idx in used_income:
                continue
            if e_sheet == i_sheet:
                continue
            if abs(e_amt - i_amt) > TOLERANCE:
                continue
            # Found a match
            used_income.add(i_idx)

            # Only warn if neither side is already marked as transfer
            if not e_txn.get("is_transfer") and not i_txn.get("is_transfer"):
                warnings.append({
                    "type": "unmarked_pair",
                    "amount": e_amt,
                    "expense": f"[{e_sheet}] [{e_txn.get('category', '')}] {e_txn.get('summary', '')}",
                    "income": f"[{i_sheet}] [{i_txn.get('category', '')}] {i_txn.get('summary', '')}",
                    "message": (
                        f"跨sheet配对: {e_sheet}支出 {e_amt:.2f} "
                        f"↔ {i_sheet}收入 {e_amt:.2f}，双方均未标记为往来"
                    ),
                })
            break

    return warnings


def _build_fx_loss_pairs(active_sheets: list[dict]) -> dict:
    """Build FX loss analysis from cross-sheet transfer pairs.

    Matches transfer expenses to transfer incomes across sheets,
    calculating the RMB difference caused by different exchange rates.

    Returns:
        {"total_loss": float, "has_loss": bool, "pairs": [FxLossPair, ...]}
    """
    expenses = []  # transfer expense entries with RMB context
    incomes = []   # transfer income entries with RMB context

    for s in active_sheets:
        rate = s["exchange_rate"]
        currency = s["currency"]
        sheet_name = s["sheet_name"]
        for t in s.get("transactions", []):
            if not t.get("is_transfer"):
                continue
            expense_amt = t.get("expense", 0)
            income_amt = t.get("income", 0)
            if expense_amt > TOLERANCE:
                expenses.append({
                    "amount": expense_amt,
                    "rmb": expense_amt * rate,
                    "currency": currency,
                    "rate": rate,
                    "sheet": sheet_name,
                    "summary": t.get("summary", ""),
                })
            if income_amt > TOLERANCE:
                incomes.append({
                    "amount": income_amt,
                    "rmb": income_amt * rate,
                    "currency": currency,
                    "rate": rate,
                    "sheet": sheet_name,
                    "summary": t.get("summary", ""),
                })

    used_expense = set()
    used_income = set()
    pairs = []

    for e_idx, exp in enumerate(expenses):
        best_idx = None
        best_diff = float("inf")

        for i_idx, inc in enumerate(incomes):
            if i_idx in used_income:
                continue
            if exp["sheet"] == inc["sheet"]:
                continue

            # Same currency → match by local amount
            # Cross currency → match by RMB equivalent
            if exp["currency"] == inc["currency"]:
                diff = abs(exp["amount"] - inc["amount"])
                tol = TOLERANCE
            else:
                diff = abs(exp["rmb"] - inc["rmb"])
                tol = max(TOLERANCE, abs(exp["rmb"]) * 0.03)

            if diff < tol and diff < best_diff:
                best_diff = diff
                best_idx = i_idx

        if best_idx is not None:
            used_expense.add(e_idx)
            used_income.add(best_idx)
            inc = incomes[best_idx]
            loss = round(inc["rmb"] - exp["rmb"], 2)
            pairs.append({
                "from_sheet": exp["sheet"],
                "to_sheet": inc["sheet"],
                "from_amount": exp["amount"],
                "from_currency": exp["currency"],
                "from_rate": exp["rate"],
                "from_rmb": round(exp["rmb"], 2),
                "to_amount": inc["amount"],
                "to_currency": inc["currency"],
                "to_rate": inc["rate"],
                "to_rmb": round(inc["rmb"], 2),
                "loss": loss,
                "summary": exp["summary"] or inc["summary"] or "内部划转",
            })

    # Collect unmatched transfers
    unmatched = []
    for e_idx, exp in enumerate(expenses):
        if e_idx not in used_expense:
            unmatched.append({
                "sheet": exp["sheet"],
                "currency": exp["currency"],
                "rate": exp["rate"],
                "amount": exp["amount"],
                "rmb": round(exp["rmb"], 2),
                "summary": exp["summary"],
                "direction": "expense",
            })
    for i_idx, inc in enumerate(incomes):
        if i_idx not in used_income:
            unmatched.append({
                "sheet": inc["sheet"],
                "currency": inc["currency"],
                "rate": inc["rate"],
                "amount": inc["amount"],
                "rmb": round(inc["rmb"], 2),
                "summary": inc["summary"],
                "direction": "income",
            })

    total_loss = round(sum(p["loss"] for p in pairs), 2)

    return {
        "total_loss": total_loss,
        "has_loss": abs(total_loss) > 1.0,
        "pairs": pairs,
        "unmatched": unmatched,
    }
